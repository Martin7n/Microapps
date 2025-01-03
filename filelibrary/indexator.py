import os
import openpyxl

def file_index(path_to_dir, all_files:bool):
    wb = openpyxl.Workbook()
    sheet = wb.active
    labels = ["Path","Full Path","Name","Size","Type","Direct link"]
    for col, label in enumerate(labels, start=1):
        sheet.cell(row=1, column=col).value = label
    wb.save('files_found.xlsx')

    if all_files is False:
        filelist = walker_docs(path_to_dir)
    else:
        filelist = walker_all(path_to_dir)
    count = 0
    for row, file in enumerate(filelist, start=2):
        sheet.cell(row=row, column=1).value = file["file"][1]
        sheet.cell(row=row, column=2).value = file["file"][2]
        sheet.cell(row=row, column=3).value = file["file"][0]
        sheet.cell(row=row, column=4).value = file["file"][3]
        sheet.cell(row=row, column=5).value = file["file"][5]
        sheet.cell(row=row, column=6).value = file["file"][2]
        sheet.cell(row=row, column=6).hyperlink = file['file'][2]
        sheet.cell(row=row, column=6).style = "Hyperlink"
        count += 1
    wb.save('files_found.xlsx')
    wb.close()
    return count


def walker_docs(path_to_dir):
    docs_and_others = ["pdf", "doc", "docx", "jpg", "jpeg", "png", "gif", "webp", "xls", "xlsx", "eml"]
    filelist = []
    for path, subdirs, files in os.walk(path_to_dir):
        for name in files:
            if name.split(".")[-1] in docs_and_others:
                found = os.path.join(path, name)
                size = os.path.getsize(found)
                created = os.path.getctime(found)
                type = f"<{name.split(".")[-1]}>"
                filelist.append({"file":[name, path, found, size, created, type]})
    return filelist

def walker_all(path_to_dir):
    filelist = []
    for path, subdirs, files in os.walk(path_to_dir):
        for name in files:
            found = os.path.join(path, name)
            size = os.path.getsize(found)
            created = os.path.getctime(found)
            type = f"<{name.split(".")[-1]}>"
            filelist.append({"file":[name, path, found, size, created, type]})
    return filelist




if __name__ == '__main__':
    print(file_index("D:\\", True))
