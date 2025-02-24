import os
import openpyxl

def file_merger(result_name, maximum_col, path, name):
    found = os.path.join(path, name)
    # print(found)
    wb_obj = openpyxl.load_workbook(found)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    wb = openpyxl.load_workbook(result_name)
    sheet = wb.active

    for row in sheet_obj.iter_rows(min_row=1, max_col=maximum_col, max_row=max_row):
        if row[0].value is not None:
            # print(row[0].value)
            inf = [x.value for x in row]
            inf.append(name)
            inf.append(found)
            sheet.append(inf)
    wb.save(result_name)
    print(f"\n \033[34m{name} merged successfully")
    return 1


def file_index(filepaths, type_obligations:False, all_types:False):
    count = 0
    collected = 0
    for path, subdirs, files in os.walk(filepaths):
        for name in files:

            if name.endswith('.xlsx'):
                if type_obligations and name.startswith('Obliga'):
                    result_name = 'obligation_found.xlsx'
                    maximum_col = 28
                    collected += file_merger(result_name, maximum_col, path, name)
                elif name.startswith('DPS_'):
                    result_name = 'dps_found.xlsx'
                    maximum_col = 32
                    collected += file_merger(result_name, maximum_col, path, name)
                elif all_types:
                    result_name = "files_found.xlsx"
                    maximum_col = 50
                    collected += file_merger(result_name, maximum_col, path, name)
                else:
                    print(f'\033[31m" skipping file.. {name}')
                count += 1

    return f"{count} files,  {collected} files collected."



if __name__ == '__main__':
    print(file_index("c:\\drob\\", False, False))
    # print(walker_xlsx("c:\\drob\\"))



