import os
from pprint import pprint
import FreeSimpleGUI as sg
import openpyxl
from openpyxl import Workbook

def file_rename(filepaths,  reg_file):

    try:

        files_in_directory = os.listdir(filepaths)
        reg_file = reg_file
        wb_obj = openpyxl.load_workbook(reg_file)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        register = {}
        for row in sheet_obj.iter_rows(min_row=2, max_col=4, max_row=max_row):
            # appendix = str(row[0].value) + row[1].value + str(row[2].value)
            appendix = str(row[1].value)
            reg = row[3].value
            register[reg] = appendix

        names = []
        for file in files_in_directory:
            names.append(file)

        count = 0


        #error log
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value = "Name"
        not_renamed = 1

        for name in names:
            search_pattern = name[:8]
            # updated for short regs
            if search_pattern[-1]=="_":
                search_pattern= search_pattern[:7]
                # print(search_pattern)
            # updated for short regs end
                if search_pattern in register:
                    new_name = register[search_pattern] + "_" + name
                    count += 1
                    # print(new_name)
                    # old_file = os.path.join(directory_path, name)
                    # new_file = os.path.join(directory_path, new_name)

                    old_file = filepaths + "//" + name
                    new_file = filepaths + "//" + new_name

                    os.rename(old_file, new_file)
                    with open("work_done.txt", "a") as file:
                        file.writelines(f"{name}, {new_name}")
            # error log
            else:
                not_renamed +=1
                sheet.cell(row=not_renamed, column=1).value = name

            wb.save('error_files.xlsx')
            wb.close()
            # error log

        return count
    except Exception:
        exit()


if __name__ == "__main__":
    pass














