import openpyxl
from vars import path_to_file

#decode and print

def decoder(excel_file_path):

    reg_file = excel_file_path
    wb_obj = openpyxl.load_workbook(reg_file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    for row in sheet_obj.iter_rows(min_row=1, max_col=1, max_row=max_row):
        print([x.value for x in row])


if __name__ == "__main__":
    decoder(path_to_file)
