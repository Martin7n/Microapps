import os

import PyPDF2
import openpyxl
from openpyxl.reader.excel import load_workbook
import re

from patterns import vin_convention_pattern, cert_number_pattern, pattern_reg, pattern_reg_short, pattern_reg_full
from vars import  path_to_file



def registry_convention(reg_file):
    '''convention part - reading the xls register'''

    column_for_appendix = 0
    column_for_reg = 0
    column_for_bro = 0


    wb_obj = openpyxl.load_workbook(reg_file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    register = {}
    for row in sheet_obj.iter_rows(min_row=2, max_col=4, max_row=max_row):
        # appendix = str(row[0].value) + row[1].value + str(row[2].value)
        # reg = row[3].value
        appendix, reg, ib = str(row[0].value), row[1].value, str(row[2].value)
        #TODO: ib usage
        register[reg] = appendix

    return register


def data_preparation(page_text, filename):
    ''' patterns: pattern_reg_full  pattern_reg_short'''
    '''   range of recognized text processed '''
    start_symbol_idx = 200
    end_symbol_idx = 5000
    additional_cutoff_idx = 100

    txt = page_text[start_symbol_idx:end_symbol_idx]

    reg_check = re.findall(pattern_reg, txt) #not in use
    reg_plate = re.findall(pattern_reg_full, txt)

    if not reg_plate:
        reg_plate = re.findall(pattern_reg_short, txt)
    certificate = re.finditer(cert_number_pattern, txt)
    vin_number = re.findall(vin_convention_pattern, txt[0:(len(txt) - additional_cutoff_idx)])
    certificate_no = "_".join([x[2] for x in certificate])


    return  [reg_plate, certificate_no, vin_number]

def ocr_file_renamer(filepaths, reg_file):

    convention_registry = registry_convention(reg_file)
    pdf_registry = {}

    filenames = os.listdir(filepaths)
    names = []

    for file_name in filenames:
        if file_name.endswith(".pdf"):
            names.append(file_name)
            pdf_file_path = filepaths + "/" + file_name
            print(pdf_file_path)

            pdfFileObj = open(pdf_file_path, 'rb')
            pdfReader = PyPDF2.PdfReader(pdfFileObj)
            pageObj = pdfReader.pages[0]
            page_text = pageObj.extract_text()

            #TODO return of the page_text and filepath

            data_prepared = data_preparation(page_text, file_name)
            pdf_registry[file_name] = data_prepared
            pdfFileObj.close()

    renaming_files(filepaths, reg_file, pdf_registry, convention_registry)

    return pdf_registry
            #
            # ''' to be excluded'''
            # txt = page_text[200:5000]
            # pdfFileObj.close()
            #
            # reg_num = re.findall(pattern_reg, txt)
            # num_reg_cert = re.findall(cert_number_pattern, txt)
            # certificate = re.finditer(cert_number_pattern, txt)
            # vin_number = re.findall(vin_convention_pattern, txt[0:(len(txt) - 100)])
            # convention_registry[pdf_file_path] = reg_num[0]
            # reg_plate = reg_num[0]
            # certificate_no = "_".join([x[2] for x in certificate])

        #TODO parse to ranaming function - 2 return the full list of recoginised data and files/paths

            # old_file = filepaths + "/" + file_name
            # new_file = filepaths + "/" + reg_num[0] + "_" + certificate_no + ".pdf"
            # if reg_plate in convention_registry:
            #     appendix = convention_registry[reg_plate]
            #     new_file = filepaths + "/" + appendix+ "_" + reg_num[0] + "_" + certificate_no + ".pdf"
            # # print(new_file, old_file)
            # os.rename(old_file, new_file)



    # return f"{filepaths},  {path_to_file}"

def renaming_files(filepaths, reg_file, pdf_registry, convention_registry):
    for filename in pdf_registry:
        reg_found = filename[0]
        vin = filename[1]
        cert = filename[2]
        old_file = filepaths + "/" + filename
        new_name = ""

        if reg_found in convention_registry:
            #adding the appendix if exists
            new_name = convention_registry[reg_found]
            #TODO split the registry as KVP and check for VIN match elif no regs.
        new_file = filepaths + "/" + new_name + "_" + reg_found + "_" + cert + "_" + ".pdf"

        print(new_file, old_file)
        # os.rename(old_file, new_file)
        with open("work_done.txt", "a") as file:
            file.writelines(f"{new_file}, {old_file}")


if __name__ == '__main__':
    pass
