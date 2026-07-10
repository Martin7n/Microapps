from datetime import datetime, date

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from esg_vehicles.brand_model import brands_t
from esg_vehicles.data_collections.headers_fields import HEADER_MAP, OUTPUT_FIELDS
from esg_vehicles.processors.data_helpers import prepare_excel_value
from esg_vehicles.models.main_class import ESGRecord
from esg_vehicles.main_esg_data_update import object_data_update
from esg_vehicles.processors.processor import process_records


def read_parse_to_class(filename):
    wb = load_workbook(filename, data_only=True)
    ws = wb.active

    headers = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    records = []

    for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            values_only=True):

        row_dict = dict(zip(headers, row))

        values = {}

        for excel_name, python_name in HEADER_MAP.items():
            values[python_name] = row_dict.get(excel_name)

        # record = ESGRecord(**values)
        row_dict = dict(zip(headers, row))
        record = ESGRecord(original=row_dict, **values)

        records.append(record)
    # print(records)
    return records




def write_records(filename, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "ESG Cleaned"

    if not records:
        return

    original_headers = list(records[0].original.keys())
    added_headers = list(OUTPUT_FIELDS.keys())
    ws.append(original_headers + added_headers)


    for record in records:
        row = []
        for header in original_headers:
            value = record.original.get(header)
            row.append(prepare_excel_value(value))
        # row = [
        #     record.original.get(header)
        #     for header in original_headers
        # ]

        # Added data
        for _, attribute in OUTPUT_FIELDS.items():

            value = getattr(record, attribute)

            # convert lists for Excel
            if isinstance(value, list):
                value = "; ".join(map(str, value))

            row.append(prepare_excel_value(value))

        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "DD.MM.YYYY"

    wb.save(filename)


def data_handler(filename, output_file, type_processing="esg_main"):
    records = read_parse_to_class(filename)

    process_records(records, type_processing)

    # try:
    #     print(records[0].brand)
    #     print(records[0].original)
    #     print(records[0].weight)
    # except Exception as e:
    #     print(e)

    write_records(output_file, records)

    return f"{len(records)} written successfully!"








def xls_parse_test(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    interval = 10
    desc = []
    car_objects = []
    object_to_write = []

    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            # max_row = interval,
                            min_col=1,
                            max_col=ws.max_column, values_only=True,
                            ):
            original_car_data = {
                # "appendix": row[0],
                # "vin": row[1],
                "brand": row[2],
                "model": row[3],
                "description": row[4],
                # "weight": row[5],
                # "measure_unit": row[6],
                # "fuel": row[7],
                # "category": row[8],
                # "emission": row[9],
                # "mileage": row[10],
                # "reg_no": row[11]
            }
            desc.append(row[0])

    data_to_write = brands_t(desc)
    # print(data_to_write)
    xls_writer(data_to_write)

    #         car_objects.append(original_car_data)
    #
    # for single_obj in car_objects:
    #     updated_object =object_data_update(single_obj)
    #     object_to_write.append(updated_object)
    #
    # return object_to_write




def xls_parse_sample(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    interval = 10

    car_objects = []
    object_to_write = []

    for row in ws.iter_rows(min_row=2,
                            max_row=ws.max_row,
                            # max_row = interval,
                            min_col=1,
                            max_col=ws.max_column, values_only=True,
                            ):
            original_car_data = {
                "appendix": row[0],
                "vin": row[1],
                "brand": row[2],
                "model": row[3],
                "description": row[4],
                "weight": row[5],
                "measure_unit": row[6],
                "fuel": row[7],
                "category": row[8],
                "emission": row[9],
                "mileage": row[10],
                "reg_no": row[11]
            }

            car_objects.append(original_car_data)

    for single_obj in car_objects:
        updated_object =object_data_update(single_obj)
        object_to_write.append(updated_object)

    return object_to_write


def xls_parse_base(file):

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    interval = 10

    car_objects = {}

    for row in ws.iter_rows(min_row=2,
            max_row=ws.max_row,
            # max_row = interval,
            min_col=1,
            max_col=ws.max_column, values_only=True,
    ):
        brand = row[1]
        model = row[2]
        emission = row[3]
        fuel = row[5]
        category = row[7]
        if brand and model and emission and fuel and category:
            mark_model = f"{brand.lower()} {str(model).lower().replace('/', '_').strip()}"
            print(mark_model)
            if mark_model not in car_objects:
                car_objects[mark_model] = []
            car_objects[mark_model].append(fuel)

    # print(car_objects)
    for k,v in car_objects.items():
        print(k, v, len(v))



def data_collection(data):
    pass


def xls_writer(data):
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Results"
    output_ws["A1"] = "Key"
    output_ws["B1"] = "Value"

    row = 2
    for key, value in data.items():
        output_ws.cell(row=row, column=1, value=key)
        output_ws.cell(row=row, column=2, value=value)
        row += 1

    output_wb.save("rs_output.xlsx")


if __name__ == '__main__':
    # xls_parse_base(r"C:\drob\ress.xlsx")
    # xls_parse_sample(r"C:\drob\sample.xlsx")
    # xls_parse_test(r"C:\drob\orf.xlsx")
    #Todo: add checks for brands and types from EAA db.
    #TODO - 2: export keywords []

    # records = read_parse_to_class(r"C:\drob\sample_fordev.xlsx")

    data_handler(r"C:\drob\sample_fordev.xlsx",
                 r"C:\drob\upadted.xlsx")

    # print(records[0].brand)
    # print(records[0].original)
    # print(records[0].weight)
